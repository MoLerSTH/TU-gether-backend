# app/api/v1/routes_events.py
from fastapi import APIRouter, HTTPException, Request, Response
from app.repositories import events_repo
from app.services import registrations_service
from app.core.session import get_session
from openpyxl import Workbook
import io
from datetime import datetime , timezone
from firebase_admin import firestore
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
from urllib.parse import quote

FACULTY_MAP = {
    "ทั้งหมด": ["ทั้งหมด", "all", ""],

    "คณะนิติศาสตร์": ["คณะนิติศาสตร์", "law", "faculty of law"],

    "คณะพาณิชยศาสตร์และการบัญชี": [
        "คณะพาณิชยศาสตร์และการบัญชี",
        "commerce",
        "accounting",
        "faculty of commerce and accountancy"
    ],

    "คณะรัฐศาสตร์": [
        "คณะรัฐศาสตร์",
        "political science",
        "faculty of political science"
    ],

    "คณะเศรษฐศาสตร์": [
        "คณะเศรษฐศาสตร์",
        "economics",
        "faculty of economics"
    ],

    "คณะสังคมสงเคราะห์ศาสตร์": [
        "คณะสังคมสงเคราะห์ศาสตร์",
        "social administration",
        "faculty of social administration"
    ],

    "คณะวารสารศาสตร์และสื่อสารมวลชน": [
        "คณะวารสารศาสตร์และสื่อสารมวลชน",
        "journalism",
        "communication"
    ],

    "คณะสังคมวิทยาและมานุษยวิทยา": [
        "คณะสังคมวิทยาและมานุษยวิทยา",
        "sociology",
        "anthropology"
    ],

    "วิทยาลัยพัฒนศาสตร์ป๋วย อึ๊งภากรณ์": [
        "วิทยาลัยพัฒนศาสตร์ป๋วย อึ๊งภากรณ์",
        "puey",
        "college of interdisciplinary studies"
    ],

    "วิทยาลัยนวัตกรรม": [
        "วิทยาลัยนวัตกรรม",
        "college of innovation"
    ],

    "วิทยาลัยสหวิทยาการ": [
        "วิทยาลัยสหวิทยาการ",
        "school of interdisciplinary studies"
    ],

    "วิทยาลัยนานาชาติปรีดีพนมยงค์": [
        "วิทยาลัยนานาชาติปรีดีพนมยงค์",
        "pridi international college"
    ],

    "คณะวิทยาการเรียนรู้และศึกษาศาสตร์": [
        "คณะวิทยาการเรียนรู้และศึกษาศาสตร์",
        "learning sciences"
    ],

    "วิทยาลัยโลกคดีศึกษา": [
        "วิทยาลัยโลกคดีศึกษา",
        "world literature college"
    ],

    "คณะศิลปศาสตร์": [
        "คณะศิลปศาสตร์",
        "faculty of liberal arts"
    ],

    "คณะศิลปกรรมศาสตร์": [
        "คณะศิลปกรรมศาสตร์",
        "fine arts"
    ],

    "คณะวิทยาศาสตร์และเทคโนโลยี": [
        "คณะวิทยาศาสตร์และเทคโนโลยี",
        "science and technology"
    ],

    "คณะวิศวกรรมศาสตร์": [
        "คณะวิศวกรรมศาสตร์",
        "engineering",
        "engineer",
        "faculty of engineering"
    ],

    "คณะสถาปัตยกรรมศาสตร์และการผังเมือง": [
        "คณะสถาปัตยกรรมศาสตร์และการผังเมือง",
        "architecture",
        "urban planning"
    ],

    "สถาบันเทคโนโลยีนานาชาติสิรินธร (SIIT)": [
        "สถาบันเทคโนโลยีนานาชาติสิรินธร (siit)",
        "siit"
    ],

    "คณะแพทยศาสตร์": [
        "คณะแพทยศาสตร์",
        "medicine",
        "medical school"
    ],

    "คณะทันตแพทยศาสตร์": [
        "คณะทันตแพทยศาสตร์",
        "dentistry"
    ],

    "คณะสหเวชศาสตร์": [
        "คณะสหเวชศาสตร์",
        "allied health"
    ],

    "คณะพยาบาลศาสตร์": [
        "คณะพยาบาลศาสตร์",
        "nursing",
        "nurse",
    ],

    "คณะสาธารณสุขศาสตร์": [
        "คณะสาธารณสุขศาสตร์",
        "public health"
    ],

    "คณะเภสัชศาสตร์": [
        "คณะเภสัชศาสตร์",
        "pharmacy"
    ],

    "วิทยาลัยแพทยศาสตร์นานาชาติจุฬาภรณ์": [
        "วิทยาลัยแพทยศาสตร์นานาชาติจุฬาภรณ์",
        "college of medical science chulabhorn"
    ]
}

router = APIRouter(tags=["events"])

def fmt_dt(v):
    """คืนค่าเวลาที่อ่านง่าย และไม่มี timezone สำหรับใส่ Excel"""
    if v is None:
        return "-"
    # Firestore Timestamp object (มี seconds/nanoseconds)
    if hasattr(v, "seconds") and hasattr(v, "nanoseconds"):
        d = datetime.fromtimestamp(v.seconds + v.nanoseconds/1e9, tz=timezone.utc)
        return d.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
    # dict รูปแบบ {"seconds": ...}
    if isinstance(v, dict) and "seconds" in v:
        d = datetime.fromtimestamp(v["seconds"], tz=timezone.utc)
        return d.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
    # datetime ปกติ (timezone-aware หรือไม่ก็ตาม)
    if isinstance(v, datetime):
        return v.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
    # อื่น ๆ แปลงเป็นสตริง
    return str(v)

def flatten_dict(d, prefix=""):
    out = {}
    if not isinstance(d, dict):
        return out
    for k, v in d.items():
        key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict):
            out.update(flatten_dict(v, key))
        else:
            out[key] = v
    return out

def get_any(data: dict, keys, default="-"):
    """ดึงค่าตามลำดับชื่อฟิลด์แรกที่เจอ"""
    for k in keys:
        # รองรับ key แบบ 'a.b.c'
        cur = data
        ok = True
        for part in str(k).split('.'):
            if isinstance(cur, dict) and part in cur:
                cur = cur[part]
            else:
                ok = False
                break
        if ok and cur not in (None, ""):
            return cur
    return default

def _to_naive(dt):
    # Excel ไม่รองรับ tzinfo
    if hasattr(dt, "replace"):
        try:
            return dt.replace(tzinfo=None)
        except Exception:
            pass
    return dt

def _get_user_by_userid(db, user_id: str):
    user_id = str(user_id)
    qs = db.collection("users").where("user_id", "==", user_id).limit(1).stream()
    for doc in qs:
        return doc.to_dict()
    return None

def _get_doc(db, col, doc_id):
    try:
        ref = db.collection(col).document(str(doc_id))
        snap = ref.get()
        return snap.to_dict() if snap.exists else None
    except Exception:
        return None
    
@router.get("/events/{event_id}/export")
async def export_event_excel(event_id: str):
    db = firestore.client()

    # 1) อ่าน registrations ของ event
    regs_ref = (
        db.collection("events")
          .document(event_id)
          .collection("registrations")
    )
    regs = [r.to_dict() for r in regs_ref.stream()]
    print(f"Found {len(regs)} registrations for event {event_id}")
    print(regs[:3])  # แสดงตัวอย่าง 3 รายการแรก
    if not regs:
        raise HTTPException(404, "ไม่พบนักศึกษาหรือผู้เข้าร่วมในกิจกรรมนี้")

    # 2) เตรียมแบ่งกลุ่ม
    students_rows = []
    publics_rows  = []

    # ---------- เริ่ม loop ผู้ลงทะเบียน ----------
    for r in regs:
        # ข้าม admin
        if (r.get("role") or "").lower() == "admin":
            continue

        user_id = r.get("user_id") or ""

        # ดึงเอกสารจาก Firestore
        sdoc = _get_doc(db, "Student", user_id)       # Student ใช้ doc id เป็นรหัสนิสิตได้
        udoc = _get_user_by_userid(db, user_id)       # users ต้อง where user_id

        # ---- จัดกลุ่มโดยดูจาก role เป็นหลัก แล้วค่อย enrich จาก Student/users ----
        def full_name_from(d):
            if not d:
                return "-"
            return (
                d.get("full_name")
                or (d.get("firstname", "") + " " + d.get("lastname", "")).strip()
                or d.get("username")
                or "-"
            )

        is_student = str(r.get("role", "")).lower() == "student"
        reg_at = _to_naive(r.get("registered_at"))

        if is_student:
            # ใช้ข้อมูลจาก Student ก่อน ถ้าไม่มีค่อย fallback ไป users / reg
            info = sdoc or udoc or {}
            students_rows.append({
                "sid":     str(user_id) if user_id else "-",
                "name":    full_name_from(info) or r.get("full_name") or "-",
                "faculty": info.get("faculty", "-"),
                "major":   info.get("major", "-"),
                "year":    info.get("grade") or info.get("year") or "-",
                "email":   info.get("email", "-"),
                "phone":   info.get("phone_num", "-"),
                "reg_at":  reg_at,
            })
        else:
            # บุคคลภายนอก (หรือ users ปกติ)
            info = udoc or sdoc or {}
            publics_rows.append({
                "name":   full_name_from(info) or r.get("full_name") or "-",
                "email":  info.get("email", "-"),
                "phone":  info.get("phone_num", "-"),
                "reg_at": reg_at,
            })

    # 3) เขียน Excel เป็น 2 บล็อค
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    # จัดสไตล์พื้นฐาน
    header_fill = PatternFill("solid", fgColor="1F4E79")  # น้ำเงินกรม
    header_font = Font(color="FFFFFF", bold=True)
    sub_header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ความกว้างคอลัมน์
    widths = {
        "A": 8,  # ลำดับ
        "B": 18, # รหัส / ชื่อ
        "C": 26, # ชื่อ-นามสกุล / คณะ
        "D": 18, # คณะ / สาขา
        "E": 18, # สาขา / ชั้นปี
        "F": 10, # ชั้นปี
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ===== ส่วนที่ 1: นักศึกษา =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="นักศึกษา")
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).alignment = center
    row += 1

    headers_std = ["ลำดับ","รหัสนักศึกษา","ชื่อ-นามสกุล","คณะ","สาขา","ชั้นปี"]
    for c, h in enumerate(headers_std, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = sub_header_font
        cell.alignment = center
        cell.border = border
    row += 1

    for i, s in enumerate(students_rows, start=1):
        values = [
            i,
            s["sid"],
            s["name"],
            s["faculty"],
            s["major"],
            s["year"],
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
        row += 1

    # เว้น 1 บรรทัด
    row += 1

    # ===== ส่วนที่ 2: บุคคลภายนอก =====
    # ใช้คอลัมน์แค่ A..C (ตามตัวอย่างในภาพใช้ A..B ก็ได้ ถ้าต้องการสั้น)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="บุคคลภายนอก")
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).alignment = center
    row += 1

    headers_pub = ["ลำดับ","ชื่อ-นามสกุล"]
    for c, h in enumerate(headers_pub, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = sub_header_font
        cell.alignment = center
        cell.border = border
    row += 1

    for i, p in enumerate(publics_rows, start=1):
        values = [i, p["name"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
        row += 1
    
     # 4) ส่งไฟล์ออก
    def make_filename(title: str, event_id: str) -> str:
        # ตัดเฉพาะอักขระต้องห้ามของชื่อไฟล์ Windows แต่ "ภาษาไทย" ให้คงไว้
        clean = re.sub(r'[\\/:*?"<>|]+', ' ', (title or '').strip())
        if not clean:
            clean = f"event_{event_id}"
        return f"{clean}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    # ดึง title ของกิจกรรม
    event_doc = db.collection("events").document(event_id).get()
    event_title = (event_doc.to_dict() or {}).get("title", f"event_{event_id}")

    utf8_name = make_filename(event_title, event_id)                       # ชื่อไฟล์จริง (ไทยได้)
    ascii_fallback = f"event_{event_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"  # สำรองเป็นอังกฤษ
    disposition = (
        f'attachment; filename="{ascii_fallback}"; '
        f"filename*=UTF-8''{quote(utf8_name)}"
    )

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return Response(
        content=mem.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    
    )


@router.get("/events")
def list_events():
    return events_repo.list_events()

@router.get("/events/{event_id}")
def get_event(event_id: str):
    try:
        return events_repo.get_event(event_id)
    except KeyError:
        raise HTTPException(404, "Event not found")
    
@router.post("/events/{event_id}/register")
async def register_event(event_id: str, request: Request):
    db = firestore.client()

    # ---- Session ----
    sess = get_session(request)
    if not sess:
        raise HTTPException(401, "กรุณาเข้าสู่ระบบก่อนลงทะเบียน")

    user_id = str(sess.get("user_id"))
    role = (sess.get("role") or "").lower()

    # ---- Normalize role ----
    if role not in ("student", "public", "admin"):
        is_student = db.collection("Student").document(user_id).get().exists
        role = "student" if is_student else "public"

    # ---- ดึงข้อมูลกิจกรรม ----
    ev_ref = db.collection("events").document(event_id)
    ev_doc = ev_ref.get()
    if not ev_doc.exists:
        raise HTTPException(404, "ไม่พบกิจกรรม")
    event = ev_doc.to_dict()

    # ---- ดึง student profile ----
    student = None
    if role == "student":
        sdoc = db.collection("Student").document(user_id).get()
        if sdoc.exists:
            student = sdoc.to_dict()

    # ---- ตรวจสอบ audience ----
    audience = normalize_text(event.get("audience"))
    if audience == "student" and role != "student":
        raise HTTPException(403, "กิจกรรมนี้เปิดรับเฉพาะนักศึกษาเท่านั้น")

    if audience == "public" and role == "student":
        raise HTTPException(403, "กิจกรรมนี้เปิดรับเฉพาะบุคคลทั่วไปเท่านั้น")

    # ---- ตรวจสอบเฉพาะนักศึกษา ----
    if student:

        # event data
        ev_fac   = normalize_faculty_auto(event.get("faculty"))
        stu_fac  = normalize_faculty_auto(student.get("faculty"))
        ev_year  = normalize_year(event.get("student_year"))

        # student data
        ev_major = normalize_major_auto(event.get("major"))
        stu_major = normalize_major_auto(student.get("major"))
        stu_year  = normalize_year(student.get("grade"))

        # คณะ
        if ev_fac not in ["", "all"] and ev_fac != stu_fac:
            raise HTTPException(403, f"กิจกรรมนี้รับเฉพาะคณะ {ev_fac} (คุณอยู่ {stu_fac})")

        # สาขา
        if ev_major not in ["", "all"] and ev_major != stu_major:
            raise HTTPException(403, f"กิจกรรมนี้รับเฉพาะสาขา {ev_major} (คุณอยู่ {stu_major})")

        # ปี
        if ev_year not in ["all", ""] and ev_year != stu_year:
            raise HTTPException(403, f"กิจกรรมนี้รับเฉพาะปี {ev_year} (คุณปี {stu_year})")

    # ---- ลงทะเบียน ----
    result = registrations_service.register(
        event_id,
        user_id,
        role=role,
    )

    return {"message": "ลงทะเบียนสำเร็จ", "data": result}



# ==========================================
# 🔥 Normalization Helpers
# ==========================================
def normalize_text(s):
    if not s:
        return ""
    return str(s).strip().lower()



def normalize_faculty_auto(text: str):
    if not text:
        return ""

    t = text.strip().lower()

    # ดึงชื่อคณะจริงแบบไทย (คณะทั้งหมดตามสกีมา)
    FACULTY_LIST = [
        "คณะนิติศาสตร์",
        "คณะพาณิชยศาสตร์และการบัญชี",
        "คณะรัฐศาสตร์",
        "คณะเศรษฐศาสตร์",
        "คณะสังคมสงเคราะห์ศาสตร์",
        "คณะวารสารศาสตร์และสื่อสารมวลชน",
        "คณะสังคมวิทยาและมานุษยวิทยา",
        "วิทยาลัยพัฒนศาสตร์ป๋วย อึ๊งภากรณ์",
        "วิทยาลัยนวัตกรรม",
        "วิทยาลัยสหวิทยาการ",
        "วิทยาลัยนานาชาติปรีดีพนมยงค์",
        "คณะวิทยาการเรียนรู้และศึกษาศาสตร์",
        "วิทยาลัยโลกคดีศึกษา",
        "คณะศิลปศาสตร์",
        "คณะศิลปกรรมศาสตร์",
        "คณะวิทยาศาสตร์และเทคโนโลยี",
        "คณะวิศวกรรมศาสตร์",
        "คณะสถาปัตยกรรมศาสตร์และการผังเมือง",
        "สถาบันเทคโนโลยีนานาชาติสิรินธร (siit)",
        "คณะแพทยศาสตร์",
        "คณะทันตแพทยศาสตร์",
        "คณะสหเวชศาสตร์",
        "คณะพยาบาลศาสตร์",
        "คณะสาธารณสุขศาสตร์",
        "คณะเภสัชศาสตร์",
        "วิทยาลัยแพทยศาสตร์นานาชาติจุฬาภรณ์",
    ]

    # 1) ถ้าเจอคำไทยตรง ให้ match ทันที
    for fac in FACULTY_LIST:
        if fac.replace(" ", "") in t.replace(" ", ""):
            return fac

    # 2) Mapping ภาษาอังกฤษขั้นต่ำที่จำเป็น
    ENG_MAP = {
        "engineering": "คณะวิศวกรรมศาสตร์",
        "engineer": "คณะวิศวกรรมศาสตร์",
        "nursing": "คณะพยาบาลศาสตร์",
        "nurse": "คณะพยาบาลศาสตร์",
        "science": "คณะวิทยาศาสตร์และเทคโนโลยี",
        "law": "คณะนิติศาสตร์",
        "medicine": "คณะแพทยศาสตร์",
        "pharmacy": "คณะเภสัชศาสตร์",
        "dentistry": "คณะทันตแพทยศาสตร์",
        "public health": "คณะสาธารณสุขศาสตร์",
    }

    for key, val in ENG_MAP.items():
        if key in t:
            return val

    return text  # ถ้ายังไม่เจอ ให้คืนค่าเดิม

def normalize_major_auto(text: str):
    if not text:
        return ""

    t = text.strip().lower()

    # โหลดรายชื่อสาขาจาก frontend (majorsData)
    # แต่ backend เรา hardcode ไว้เป็น list เพื่อความเร็ว
    ALL_MAJORS = []

    # 1) ดึงสาขาจาก majorsData.js (ยัดทีเดียวทั้ง object)
    #    คุณต้อง copy รายชื่อทั้งหมดของ majorsData ใส่ ALL_MAJORS
    #    ผมจะเตรียมให้แบบ auto ด้านล่าง

    # ---- ตรวจ exact match ไทย ----
    for m in ALL_MAJORS:
        if m.replace(" ", "") == t.replace(" ", ""):
            return m

    # ---- ตรวจคำยาว เช่น "วิศวกรรมซอฟต์แวร์" อยู่ใน "วิศวกรรมศาสตรบัณฑิต สาขาวิชาวิศวกรรมซอฟต์แวร์" ----
    for m in ALL_MAJORS:
        if m.replace(" ", "") in t.replace(" ", ""):
            return m

    # ---- ตรวจภาษาอังกฤษขั้นต่ำ ----
    ENG_MAP = {
        "software engineering": "วิศวกรรมซอฟต์แวร์",
        "computer engineering": "วิศวกรรมคอมพิวเตอร์",
        "civil engineering": "วิศวกรรมโยธา",
        "mechanical engineering": "วิศวกรรมเครื่องกล",
        "electrical engineering": "วิศวกรรมไฟฟ้า",
    }

    for key, val in ENG_MAP.items():
        if key in t:
            return val

    return text

def normalize_year(y):
    y = normalize_text(y)
    if y in ["all", "ทั้งหมด", ""]:
        return "all"
    try:
        return int(y)
    except:
        return "all"

@router.delete("/events/{event_id}/register")
def unregister_event(event_id: str, request: Request):
    sess = get_session(request)
    if not sess:
        raise HTTPException(401, "Not logged in")
    try:
        return registrations_service.unregister(event_id, sess["user_id"])
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Internal error: {e}")

@router.get("/events/{event_id}/register/status")
def registration_status(event_id: str, request: Request):
    sess = get_session(request)
    if not sess:
        raise HTTPException(401, "Not logged in")
    try:
        return registrations_service.status(event_id, sess["user_id"])
    except Exception as e:
        raise HTTPException(500, f"Internal error: {e}")
    

